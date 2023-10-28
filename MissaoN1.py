
from typing import List, Tuple, Union
import customtkinter as ctk
from tkinter import*
from tkinter import messagebox
import openpyxl, xlrd
import pathlib
from openpyxl import Workbook
import csv
import pandas as pd
import numpy as np

#Janela Principal
janela = ctk.CTk()
janela.title("Gerenciador de Acesso a Perfis e Sistemas")
janela.geometry("800x450")
janela.maxsize(width=800,height=450)
janela.minsize(width=500,height=300)
janela.resizable(width=False,height=False)

janela._set_appearance_mode("#222B35")



#Novas Janelas
menu = ctk.CTkTabview(janela,width=800,height=450,fg_color="#222B35",bg_color="#1d2229",corner_radius=5,border_color="#898d93")
menu.pack()
menu.add("Cadastro dos Sistemas")
menu.add("Cadastro dos Perfis")
menu.add("Cadastro da Matriz SOD")
menu.add("Consulta")




ficheiro = pathlib.Path("gestao.xlsx")
        
#Text Variables  - Variáveis de Texto
nome_sistema_value = StringVar()
codigo_sistema_value = StringVar()

codigo_sistema_perfil_value = StringVar()
nome_perfil_value = StringVar()
#descricao_perfil_value = StringVar()

#Entrys e Labels - Caixas de Texto e Caixas de entrada de Dados

#Tela Cadastro de Sistemas
en_nome_sistema_cadastrado= ctk.CTkEntry(menu.tab("Cadastro dos Sistemas"),textvariable=nome_sistema_value,width=300, font=("Century Gothic bold",16)).place(x=350,y=60)
en_codigo_sistema_cadastrado= ctk.CTkEntry(menu.tab("Cadastro dos Sistemas"),textvariable=codigo_sistema_value,width=300, font=("Century Gothic bold",16)).place(x=10,y=60)

lb_nome_sistema_cadastrado = ctk.CTkLabel(menu.tab("Cadastro dos Sistemas"),text="Nome do Sistema", font=("Century Gothic bold", 16),text_color="#fff").place(x=350,y=30)
lb_codigo_sistema_cadastrado = ctk.CTkLabel(menu.tab("Cadastro dos Sistemas"),text="Código do Sistema", font=("Century Gothic bold", 16),text_color="#fff").place(x=10,y=30)       

#Tela Cadastro do Perfis
en_codigo_sistema_cadastrado_p= ctk.CTkEntry(menu.tab("Cadastro dos Perfis"),textvariable=codigo_sistema_perfil_value,width=300, font=("Century Gothic bold",16)).place(x=10,y=60)
en_nome_perfil_cadastrado= ctk.CTkEntry(menu.tab("Cadastro dos Perfis"),textvariable=nome_perfil_value,width=300, font=("Century Gothic bold",16),text_color="#fff").place(x=350,y=60)
en_descricao_perfil= ctk.CTkTextbox(menu.tab("Cadastro dos Perfis"),width=640,height=70).place(x=10,y=150)    

lb_codigo_sistema_perfil = ctk.CTkLabel(menu.tab("Cadastro dos Perfis"),text="Código do Sistema", font=("Century Gothic bold", 16),text_color="#fff").place(x=10,y=30)
lb_nome_perfil = ctk.CTkLabel(menu.tab("Cadastro dos Perfis"),text="Nome do Perfil", font=("Century Gothic bold", 16),text_color="#fff").place(x=350,y=30) 
lb_descricao_perfil = ctk.CTkLabel(menu.tab("Cadastro dos Perfis"),text="Descrição Detalhada do Perfil", font=("Century Gothic bold", 16),text_color="#fff").place(x=10,y=100)


        
def submit_cadastro_sistemas():
     
    #Pegando os dados das entrys
    nome_sistema = nome_sistema_value.get()
    codigo_sistema = codigo_sistema_value.get()
    codigo_sistema_perfil = codigo_sistema_perfil_value.get()
    nome_perfil = nome_perfil_value.get()
    

            

    
    if(nome_sistema =="") or (codigo_sistema ==""): #Validar campos vazios
        messagebox.showerror("Systema", "ERRO\nPor gentileza preencha todos os campos para prosseguir! ")
    else:
        #codigo_sistema = codigo_sistema+1
        ficheiro  = openpyxl.load_workbook("gestao.xlsx")
                
        folha1 = ficheiro.get_sheet_by_name("Cadastro_Sistema")
        folha1.cell(column =1,row = folha1.max_row+1, value = nome_sistema)
        folha1.cell(column =2,row = folha1.max_row, value = codigo_sistema)
        
        ficheiro.save(r"gestao.xlsx")
        messagebox.showinfo("Systema","Dados salvos com sucesso!")

def submit_cadastro_perfil():
     
    #Pegando os dados das entrys
    nome_sistema = nome_sistema_value.get()
    codigo_sistema = codigo_sistema_value.get()
    codigo_sistema_perfil = codigo_sistema_perfil_value.get()
    nome_perfil = nome_perfil_value.get()
    descricao_perfil = en_descricao_perfil.get(0.0,END)
    #No caso das observações é necesário passar o início e fim de captura de dados.
    #(0.0 indica que é da linha 0, coluna 0 e END que é até o fim)
            

    #Validar campos vazios
    if(nome_sistema ==0):
        messagebox.showerror("Systema", "ERRO\nPor gentileza preencha todos os campos para prosseguir! ")
    else:
        #codigo_sistema = codigo_sistema+1
        ficheiro  = openpyxl.load_workbook("gestao.xlsx")
                
        

        folha2 = ficheiro.get_sheet_by_name("Cadastro_Perfil")
        folha2.cell(column =1,row = folha2.max_row+1, value = codigo_sistema_perfil)
        folha2.cell(column =2,row = folha2.max_row, value = nome_perfil)
        #folha2.cell(column =3,row = folha2.max_row, value  = descricao_perfil)
        

                
        ficheiro.save(r"gestao.xlsx")
        messagebox.showinfo("Systema","Dados salvos com sucesso!")

        df_sistemas = pd.read_excel('gestao.xlsx',sheet_name="Cadastro_Sistema")
                
        #df_nome_sistemas = df_sistemas["Nome do Sistema"]
        lista_sistemas = np.array(df_sistemas["Nome do Sistema"].to_list())
                
        
    
            
#btn_clear = ctk.CTkButton(janela, text = "Limpar dados".upper(),command = submit_cadastro_perfil,fg_color = "#3cb371",hover_color =  "#3a865b").place(x=500,y=420)
btn_submit_sistema = ctk.CTkButton(menu.tab("Cadastro dos Perfis"),text = "Salvar dados".upper(),command=submit_cadastro_perfil,fg_color = "#008000",hover_color =  "#3a865b").place(x=320,y=250)
btn_submit_perfil = ctk.CTkButton(menu.tab("Cadastro dos Sistemas"),text = "Salvar dados".upper(),command=submit_cadastro_sistemas,fg_color = "#008000",hover_color =  "#3a865b").place(x=320,y=250)


#Listas de Consulta de Sistemas       
        
df_sistemas = pd.read_excel('gestao.xlsx',sheet_name="Cadastro_Sistema")
lista_sistemas = np.array(df_sistemas["Nome do Sistema"].to_list())
lista_codigo_sistemas = np.array(df_sistemas["Código do Sistema"].to_list())
#codigo_sistema = 1
#cprint(f'Lista Sistemas:{lista_sistemas}\n')
print(f'Lista Código sistemas{lista_codigo_sistemas}')
#print(f'Novo código{novo_codigo}')
    
        
Combo_Sistemas = ctk.CTkComboBox(menu.tab("Consulta"),values=lista_sistemas,width=250,height=15,state="normal").place(x=10,y=60)
       
    
        

   
    
#Criando os botões de submeter e limpar dados (Obs para funcionar as funções precisam estar criadas)
#Submeter
#btn_submit = ctk.CTkButton(menu.tab("Cadastro dos Sistemas"),text = "Salvar dados".upper(),command = submit_cadastro_perfil,fg_color="#151",hover_color="#131").place(x=200,y=220)
        

janela.mainloop()